import openpyxl

# Open a local excel work book
wb = openpyxl.Workbook()

sheetName =  wb.get_sheet_names()

sheet = wb.get_sheet_by_name(sheetName[0])

sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('.\\Documents\\myExample.xlsx')

# To open a existing workbook use
# openpyxl.load_workbook('')

sheet2 = wb.create_sheet() # Creates new work sheet

# Change Sheet name
sheet2.title = 'My New Sheet Name'
wb.save('.\\Documents\\myExample2.xlsx')

# When creating sheet change position of the sheet
wb.create_sheet(index = 0, title = 'My Other Sheet')
wb.save('.\\Documents\\myExample3.xlsx')
